import os
import pandas as pd
from glob import glob
from openpyxl.utils import get_column_letter
import openpyxl
import xlwt
from openpyxl import load_workbook


def convert_xlsx_to_xls(xlsx_file, xls_file):
    wb_xlsx = load_workbook(filename=xlsx_file)

    wb_xls = xlwt.Workbook()

    for sheet_name in wb_xlsx.sheetnames:
        ws_xlsx = wb_xlsx[sheet_name]
        ws_xls = wb_xls.add_sheet(sheet_name)

        for row in ws_xlsx.iter_rows():
            for cell in row:
                ws_xls.write(cell.row - 1, cell.column - 1, cell.value)

    wb_xls.save(xls_file)
    print(f"Файл успешно сконвертирован и сохранен как {xls_file}")


data_to_parse = [
    ['Продавец:', '(2)'],
    ['ИНН/КПП продавца:', '(2б)'],
    ['Документ об отгрузке', '(5а)'],
]

data_to_parse_no_index = [
    ['Продавец', '(2)'],
    ['ИНН/КПП продавца', '(2б)'],
    ['Документ об отгрузке:', '(5а)'],
]

target_headers = [
    "А", "1", "1а", "1б", "2", "2а", "3", "4", "5", "6", "7", "8", "9", "10", "10а", "11", "12", "12а", "13", "14", "(5а)", "(2)", "(2б)"
]


def merge_xlsx_by_headers(source_path, target_path):
    source_wb = openpyxl.load_workbook(source_path)
    target_wb = openpyxl.load_workbook(target_path)

    source_sheet = source_wb.active
    target_sheet = target_wb.active

    target_headers = []
    for cell in target_sheet[1]:
        target_headers.append(cell.value)

    source_headers = []
    for cell in source_sheet[1]:
        source_headers.append(cell.value)

    column_mapping = {}
    for idx, header in enumerate(target_headers, 1):
        if header in source_headers:
            source_col = source_headers.index(header) + 1
            column_mapping[idx] = source_col
        else:
            column_mapping[idx] = None

    last_target_row = target_sheet.max_row

    for source_row in range(2, source_sheet.max_row + 1):
        last_target_row += 1
        for target_col, source_col in column_mapping.items():
            if source_col is not None:
                cell_value = source_sheet.cell(row=source_row, column=source_col).value
                target_sheet.cell(row=last_target_row, column=target_col, value=cell_value)

    target_wb.save(target_path)
    print(f"Данные успешно объединены в файл: {target_path}")


def parse_xls_xlsx_get_data(file_path, data_to_get):
    try:
        if file_path.lower().endswith('.xls'):
            df = pd.read_excel(file_path, header=None, engine='xlrd')
        elif file_path.lower().endswith('xlsx'):
            df = pd.read_excel(file_path, header=None, engine='openpyxl')
        df = df.fillna('')  # NaN
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return

    found_rows = []
    for index, row in df.iterrows():
        row_str = '|'.join(row.astype(str)).lower()

        if data_to_get[0].lower() in row_str:
            non_empty_cells = [file_path] + [cell for cell in row if str(cell).strip() not in ('', 'nan')]
            found_rows.append(non_empty_cells)

    return found_rows


def save_to_excel(data_df, output_file="результат.xlsx"):
    try:
        # Добавляем новый столбец с очищенным номером
        if 'А' in data_df.columns:
            # Создаем новый столбец, удаляя '@' и '-' из значений столбца 'А'
            data_df.insert(1, '(номер без @ и без -)',
                           data_df['А'].astype(str).str.replace('[@-]', '', regex=True))

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data_df.to_excel(
                writer,
                index=False,
                sheet_name='Данные',
            )
            worksheet = writer.sheets['Данные']
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

        print(f"✅ Таблица успешно сохранена в файл: {output_file}")

    except Exception as e:
        print(f"❌ Ошибка при сохранении: {e}")


def find_and_extract_tables(file_path):
    print(f"Обработка файла: {file_path}")

    if file_path.lower().endswith('.xls'):
        df_list = [pd.read_excel(file_path, header=None, engine='xlrd')]
    elif file_path.lower().endswith('xlsx'):
        # Для xlsx читаем все листы
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        df_list = [pd.read_excel(xls, sheet_name=sheet, header=None)
                   for sheet in xls.sheet_names]

    all_tables = []

    for df in df_list:
        df = df.fillna('')
        tables_in_sheet = []
        current_table_start = None

        for row_idx in range(len(df)):
            row_values = [str(cell) for cell in df.iloc[row_idx].values if str(cell).strip() not in ('', 'nan')]

            if all(header in row_values for header in target_headers[:6]):
                if current_table_start is not None:
                    tables_in_sheet.append((current_table_start, row_idx - 1))
                current_table_start = row_idx
            elif current_table_start is not None and len(row_values) == 0:
                tables_in_sheet.append((current_table_start, row_idx - 1))
                current_table_start = None

        if current_table_start is not None:
            tables_in_sheet.append((current_table_start, len(df) - 1))

        for start, end in tables_in_sheet:
            try:
                data_df = pd.read_excel(
                    file_path,
                    header=start,
                    usecols=lambda x: str(x) in target_headers,
                    nrows=end - start,
                    engine='openpyxl' if file_path.lower().endswith('xlsx') else 'xlrd'
                )

                data_df = data_df.dropna(how='all')

                for i in range(len(data_to_parse)):
                    to_add = parse_xls_xlsx_get_data(file_path, data_to_parse[i])
                    if to_add and len(to_add[0]) >= 4:
                        data_df[to_add[0][3]] = to_add[0][2]
                    else:
                        to_add1 = parse_xls_xlsx_get_data(file_path, data_to_parse_no_index[i])
                        if 'тот' in to_add1[0][2]:
                            'Счет-фактура  № ЦБ-3616 от 21 июля 2025 г.'
                            to_add2 = parse_xls_xlsx_get_data(file_path, 'Счет-фактура')
                            data_df[data_to_parse_no_index[i][1]] = to_add2[0][2]
                        else:
                            data_df[data_to_parse_no_index[i][1]] = to_add1[0][2]

                all_tables.append(data_df)

            except Exception as e:
                print(f"Ошибка при обработке таблицы (строки {start}-{end}): {e}")

    return all_tables


if __name__ == "__main__":
    folder_path = "upd"
    target_path = "all_data_file.xlsx"
    temp_file = "temp_data_file.xlsx"

    excel_files = glob(os.path.join(folder_path, "*.xls*"))

    new_path = input(f'Введите путь к папке с файлами XLSX и XLS (стандартно - это папка "{folder_path}"): ')
    if new_path:
        folder_path = new_path
    new_path = input("Введите путь к файлу all_data_file.xlsx (стандартно - это текущая папка): ")
    if new_path:
        target_path = os.path.join(new_path, "all_data_file.xlsx")

    if not os.path.exists(target_path):
        # Создаем DataFrame с нужными колонками, включая новую колонку
        initial_columns = target_headers.copy()
        initial_columns.insert(1, '(номер без @ и без -)')
        pd.DataFrame(columns=initial_columns).to_excel(target_path, index=False)

    for file in excel_files:
        tables = find_and_extract_tables(file)
        for i, table in enumerate(tables, 1):
            print(f"Обработка таблицы {i} из файла {file}")
            save_to_excel(table, temp_file)
            merge_xlsx_by_headers(temp_file, target_path)

    if os.path.exists(temp_file):
        os.remove(temp_file)

    print("Обработка всех файлов завершена!")

    convert_xlsx_to_xls(target_path, target_path[:-1])

"""
Linux and Mac:
Terminal
pip install openpyxl xlrd xlwt pandas

Windows:
CMD
pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org --cert /dev/null openpyxl
pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org --cert /dev/null xlrd
pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org --cert /dev/null xlwt
pip install --trusted-host pypi.org --trusted-host files.pythonhosted.org --cert /dev/null pandas
"""

