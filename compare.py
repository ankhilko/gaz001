import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows


def merge_and_color_excel_files(file1_path, file2_path, column_one, column_two, output_path, col_mark):
    # Загрузка данных из файлов
    if file1_path.lower().endswith('.xls'):
        df1 = [pd.read_excel(file1_path, engine='xlrd')]
    elif file1_path.lower().endswith('xlsx'):
        df1 = pd.read_excel(file1_path)


    if file2_path.lower().endswith('.xls'):
        df2 = [pd.read_excel(file2_path, engine='xlrd')]
    elif file2_path.lower().endswith('xlsx'):
        df2 = pd.read_excel(file2_path)

    # Находим строки из файла2, где столбец "один" не пустой и не NaN
    non_empty_mask = df2[column_one].notna() & (df2[column_one] != '')
    new_rows = df2[non_empty_mask].copy()

    # Добавляем метку для строк из файла2
    new_rows['_source'] = col_mark

    # Объединяем данные
    merged_df = pd.concat([df1, new_rows], ignore_index=True)

    # Сохраняем объединенные данные во временный файл
    temp_output = "temp_" + output_path
    merged_df.to_excel(temp_output, index=False)

    # Загружаем временный файл для форматирования
    wb = load_workbook(temp_output)
    ws = wb.active

    # Создаем стили
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    # Получаем список значений из столбца "два" в новых строках
    new_values = set(new_rows[column_two].dropna().astype(str))

    # Определяем индексы столбцов
    headers = [cell.value for cell in ws[1]]
    col_one_idx = headers.index(column_one) + 1
    col_two_idx = headers.index(column_two) + 1
    source_col_idx = len(headers) + 1  # Добавляем столбец с меткой после всех данных

    # Применяем форматирование
    for row_idx in range(2, ws.max_row + 1):
        is_from_file2 = ws.cell(row=row_idx, column=source_col_idx).value == col_mark
        cell_value = str(ws.cell(row=row_idx, column=col_two_idx).value)

        if cell_value in new_values:
            # Форматируем совпадающие строки из file1: зелёная заливка
            for cell in ws[row_idx][:-1]:
                cell.fill = yellow_fill

    # Удаляем временный столбец с меткой
    ws.delete_cols(source_col_idx)

    # Сохраняем результат
    wb.save(output_path)

    # Удаляем временный файл
    import os
    os.remove(temp_output)


if __name__ == "__main__":
    file1_path = "main_snab.xlsx"
    file2_path = "main_alts.xlsx"

    merge_and_color_excel_files(
        file1_path=file1_path,
        file2_path=file2_path,
        column_one="Поставщик",
        column_two="(номер без @ и без -)",
        output_path="result.xlsx",
        col_mark=file2_path
    )

