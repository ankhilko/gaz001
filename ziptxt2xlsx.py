import os
import zipfile
from openpyxl import Workbook


def process_zip_files():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    for zip_file in os.listdir(script_dir):
        if zip_file.endswith('.zip'):
            zip_path = os.path.join(script_dir, zip_file)
            zip_name = os.path.splitext(zip_file)[0]

            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    for file_info in zip_ref.infolist():
                        if file_info.filename.endswith('.txt') and not file_info.is_dir():
                            txt_name = os.path.splitext(file_info.filename)[0]

                            wb = Workbook()
                            ws = wb.active
                            ws.title = txt_name[:30]

                            with zip_ref.open(file_info) as txt_file:
                                content = txt_file.read().decode('utf-8').splitlines()

                                for line in content:
                                    if line.strip():
                                        columns = line.split('\t')
                                        ws.append(columns)

                            for col in ws.columns:
                                max_length = 0
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

                            output_path = os.path.join(script_dir, f"{zip_name}.xlsx")
                            wb.save(output_path)
                            print(f"Данные из {file_info.filename} сохранены в {output_path}")

            except zipfile.BadZipFile:
                print(f"Ошибка: файл {zip_file} не является ZIP-архивом или поврежден")


def get_column_letter(col_idx):
    letter = ''
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter


if __name__ == "__main__":
    process_zip_files()
